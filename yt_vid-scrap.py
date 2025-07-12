import requests
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

API_KEY = "AIzaSyBfpST2JsKRs36I7okPfdfNV9YEgBrRie4"

def get_video_title(video_id, api_key):
    """Gets the title of a YouTube video given its ID."""
    url = f"https://www.googleapis.com/youtube/v3/videos?part=snippet&id={video_id}&key={api_key}"
    try:
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()
        if data['items']:
            return data['items'][0]['snippet']['title']
        else:
            return "Video Title Not Found"
    except requests.exceptions.RequestException as e:
        print(f"Failed to retrieve video title for {video_id}: {e}")
        return "Video Title Not Found"
    except KeyError as e:
        print(f"Failed to parse video title for {video_id}: {e}")
        return "Video Title Not Found"

video_ids = []
while True:
    video_id = input("Enter a YouTube video ID (or type 'done'): ")
    if video_id.lower() == 'done':
        break
    video_ids.append(video_id)

all_comments = []

for video_id in video_ids:
    video_title = get_video_title(video_id, API_KEY)
    url = f"https://www.googleapis.com/youtube/v3/commentThreads?part=snippet&videoId={video_id}&key={API_KEY}&maxResults=5"  # Get top 5 comments

    try:
        response = requests.get(url)
        response.raise_for_status()

        data = response.json()
        for item in data['items']:
            comment = item['snippet']['topLevelComment']['snippet']['textDisplay']
            all_comments.append([video_title, comment])

    except requests.exceptions.RequestException as e:
        print(f"Failed to retrieve comments for video {video_title}: {e}")
    except KeyError as e:
        print(f"Failed to parse comments for video {video_title}: {e}")

df = pd.DataFrame(all_comments, columns=["Video Title", "Comment"])

file_path = "C:\\Users\\PC\\Desktop\\comments.xlsx"
df.to_excel(file_path, index=False, engine='openpyxl')


workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Define styles
header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
font = Font(name='Arial', size=10)

for cell in sheet[1]:
    cell.fill = header_fill
    cell.border = border
    cell.alignment = alignment
    cell.font = font

for row in sheet.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = alignment
        cell.font = font


for column_cells in sheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

workbook.save(file_path)